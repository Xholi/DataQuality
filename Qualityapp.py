import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import plotly.graph_objects as go
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import tempfile

# Define validation rules
def is_integer(value):
    return pd.api.types.is_integer_dtype(value)

def is_string(value):
    return pd.api.types.is_string_dtype(value)

def is_numeric_or_null(value):
    return pd.isnull(value) or pd.api.types.is_numeric_dtype(value)

def is_string_or_null(value):
    return pd.isnull(value) or pd.api.types.is_string_dtype(value)

def is_valid_mand_ind(value):
    return value in ['Y', 'N']

def is_present(value):
    return pd.notnull(value)

def is_in_predefined_set(value, predefined_set):
    return value in predefined_set

# Validation functions for each column
def validate_corp_no(row):
    return is_integer(row['CORP_NO'])

def validate_erp_no(row):
    return is_integer(row['ERP_NO'])

def validate_descr(row):
    return is_string(row['DESCR'])

def validate_property_term(row):
    return is_string(row['PROPERTY_TERM']) and is_present(row['PROPERTY_TERM'])

def validate_property_value(row):
    return is_numeric_or_null(row['PROPERTY_VALUE'])

def validate_clean_property_value(row):
    return is_numeric_or_null(row['CLEAN_PROPERTY_VALUE'])

def validate_extra_details(row):
    return is_string_or_null(row['EXTRA_DETAILS'])

def validate_prop_fft(row):
    return is_string_or_null(row['PROP_FFT'])

def validate_property_uom(row, predefined_uoms):
    return is_string(row['PROPERTY_UOM']) and is_in_predefined_set(row['PROPERTY_UOM'], predefined_uoms)

def validate_suggested_uom(row, predefined_uoms):
    return is_string(row['SUGGESTED_UOM']) and is_in_predefined_set(row['SUGGESTED_UOM'], predefined_uoms)

def validate_uom_rules(row, predefined_uom_rules):
    return is_string(row['UOM_RULES']) and is_in_predefined_set(row['UOM_RULES'], predefined_uom_rules)

def validate_data_type_rules(row, predefined_data_type_rules):
    return is_string(row['DATA_TYPE_RULES']) and is_in_predefined_set(row['DATA_TYPE_RULES'], predefined_data_type_rules)

def validate_data_type(row, predefined_data_types):
    return is_string(row['DATA_TYPE']) and is_in_predefined_set(row['DATA_TYPE'], predefined_data_types)

def validate_originating_plant_trm(row):
    return is_string(row['ORIGINATING_PLANT_TRM'])

def validate_originating_division(row):
    return is_string(row['ORIGINATING_DIVISION'])

def validate_plant_group(row):
    return is_string(row['PLANT_GROUP'])

def validate_mand_ind(row):
    return is_valid_mand_ind(row['MAND_IND']) and is_present(row['MAND_IND'])

def validate_mand_empty(row):
    return is_string_or_null(row['MAND_EMPTY'])

def validate_clean_property_uom(row):
    return is_string_or_null(row['CLEAN_PROPERTY_UOM'])

# Combined validation function
def validate_row(row, predefined_uoms, predefined_uom_rules, predefined_data_type_rules, predefined_data_types):
    return (
        validate_corp_no(row) and
        validate_erp_no(row) and
        validate_descr(row) and
        validate_property_term(row) and
        validate_property_value(row) and
        validate_clean_property_value(row) and
        validate_extra_details(row) and
        validate_prop_fft(row) and
        validate_property_uom(row, predefined_uoms) and
        validate_suggested_uom(row, predefined_uoms) and
        validate_uom_rules(row, predefined_uom_rules) and
        validate_data_type_rules(row, predefined_data_type_rules) and
        validate_data_type(row, predefined_data_types) and
        validate_originating_plant_trm(row) and
        validate_originating_division(row) and
        validate_plant_group(row) and
        validate_mand_ind(row) and
        validate_mand_empty(row) and
        validate_clean_property_uom(row)
    )

# Function to validate data
def validate_data(data):
    predefined_uoms = {'MILLIMETER', 'AMPERE', 'VOLT'}
    predefined_uom_rules = {'RULE1', 'RULE2'}
    predefined_data_type_rules = {'NUMERIC', 'STRING'}
    predefined_data_types = {'MEASURED_NUMBER'}

    valid_entries = []
    invalid_entries = []

    for index, row in data.iterrows():
        if validate_row(row, predefined_uoms, predefined_uom_rules, predefined_data_type_rules, predefined_data_types):
            valid_entries.append(row)
        else:
            invalid_entries.append(row)

    return pd.DataFrame(valid_entries), pd.DataFrame(invalid_entries)

# Function to calculate completeness based on new rules
def calculate_completeness(data):
    complete_entries = 0
    for index, row in data.iterrows():
        property_value = row['PROPERTY_VALUE']
        property_uom = row['PROPERTY_UOM']
        data_type_rules = row['DATA_TYPE_RULES']
        
        if pd.api.types.is_numeric_dtype(property_value) and pd.notnull(property_uom) and data_type_rules == 'NUMERIC':
            complete_entries += 1
        elif pd.api.types.is_string_dtype(property_value) and pd.isnull(property_uom) and data_type_rules == 'STRING':
            complete_entries += 1
    
    completeness = (complete_entries / len(data)) * 100 if len(data) > 0 else 0
    return completeness

# Function to calculate percentage of duplicated PODs
def calculate_duplicated_pod(data):
    if 'POD' in data.columns:
        total_pods = data['POD'].shape[0]
        duplicated_pods = data['POD'].duplicated().sum()
        duplicated_percentage = (duplicated_pods / total_pods) * 100 if total_pods > 0 else 0
        return duplicated_percentage
    return 0

# Function to save stats to Excel
def save_stats_to_excel(stats, filename="data_quality_stats.xlsx"):
    df = pd.DataFrame([stats])
    try:
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        writer.save()
    except FileNotFoundError:
        df.to_excel(filename, index=False)

# Function to send email with attachment
def send_email(to_email, subject, body, attachment_path):
    from_email = "your_email@example.com"  # Replace with your email
    from_password = "your_password"  # Replace with your email password

    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    attachment = open(attachment_path, "rb")

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= " + attachment_path)

    msg.attach(part)

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(from_email, from_password)
    text = msg.as_string()
    server.sendmail(from_email, to_email, text)
    server.quit()

# Streamlit app
st.title('Data Quality Check Tool')

# File upload
uploaded_file = st.file_uploader("Upload your data file (CSV)", type=["csv"])

if uploaded_file is not None:
    data = pd.read_csv(uploaded_file)

    # Validate data
    valid_data, invalid_data = validate_data(data)

    # Calculate statistics
    total_entries = len(data)
    valid_entries = len(valid_data)
    invalid_entries = len(invalid_data)
    accuracy = (valid_entries / total_entries) * 100 if total_entries > 0 else 0
    failure_rate = (invalid_entries / total_entries) * 100 if total_entries > 0 else 0
    completeness = calculate_completeness(data)
    duplicated_pod_percentage = calculate_duplicated_pod(data)

    # Save stats to Excel
    stats = {
        'Filename': uploaded_file.name,
        'Total Entries': total_entries,
        'Valid Entries': valid_entries,
        'Invalid Entries': invalid_entries,
        'Accuracy': accuracy,
        'Completeness': completeness,
        'Duplicated POD Percentage': duplicated_pod_percentage
    }
    save_stats_to_excel(stats)

    # Display statistics
    tab1, tab2 = st.tabs(["Statistics", "Data Tables"])

    with tab1:
        st.header("Statistics")
        st.write(f"Total items processed: {total_entries}")
        st.write(f"Accuracy: {accuracy:.2f}%")
        st.write(f"Valid score: {valid_entries:.2f}%")
        st.write(f"Failed validation: {failure_rate:.2f}%")
        st.write(f"Completeness: {completeness:.2f}%")
        st.write(f"Duplicated POD: {duplicated_pod_percentage:.2f}%")

        # Stats per plant and division
        if 'ORIGINATING_PLANT_TRM' in data.columns:
            st.write("Stats per plant:")
            st.write(data['ORIGINATING_PLANT_TRM'].value_counts())

        if 'ORIGINATING_DIVISION' in data.columns:
            st.write("Stats per division:")
            st.write(data['ORIGINATING_DIVISION'].value_counts())

        # Multi-line graph for accuracy and completeness over time
        try:
            historical_stats = pd.read_excel("data_quality_stats.xlsx")
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=historical_stats['Filename'], y=historical_stats['Accuracy'], mode='lines+markers', name='Accuracy'))
            fig.add_trace(go.Scatter(x=historical_stats['Filename'], y=historical_stats['Completeness'], mode='lines+markers', name='Completeness'))
            fig.update_layout(title='Accuracy and Completeness Over Time', xaxis_title='File', yaxis_title='Percentage')
            st.plotly_chart(fig)
        except FileNotFoundError:
            st.write("No historical data available to plot.")

    with tab2:
        st.header("Data Tables")

        # Valid entries table with conditional formatting
        st.subheader("Valid Entries")
        st.dataframe(valid_data.style.applymap(lambda x: 'background-color: lightgreen' if pd.notnull(x) else ''))

        # Invalid entries table
        st.subheader("Invalid Entries")
        st.dataframe(invalid_data)

        # Email sending section
        st.subheader("Send Invalid Entries via Email")
        recipient_email = st.text_input("Recipient Email")
        if st.button("Send Email"):
            if recipient_email:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp:
                    invalid_data.to_csv(tmp.name, index=False)
                    send_email(recipient_email, "Invalid Items to Investigate", "Good day", tmp.name)
                st.success("Email sent successfully!")
            else:
                st.error("Please enter a recipient email address.")
